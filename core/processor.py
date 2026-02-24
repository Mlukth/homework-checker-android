# "core/processor.py"
# core/processor.py
import os
import pandas as pd
from typing import Dict, List, Callable, Optional
from .file_renamer import FileRenamer

class HomeworkProcessor:
    def __init__(self):
        self.file_renamer = FileRenamer()

    def process_homework(self, roster_path: str, homework_dir: str, output_dir: str, 
                        rename_format: dict, log_callback: Optional[Callable] = None):
        """
        主处理函数
        """
        try:
            project_name = os.path.basename(homework_dir.rstrip(os.sep)).upper()

            self._log(f"\n{'='*50}", log_callback)
            self._log(f"处理 {project_name} 项目", log_callback)
            self._log(f"{'='*50}\n", log_callback)

            # 读取花名册
            df = self._read_roster(roster_path)
            all_students = set(df['姓名'].tolist())
            id_to_name = {str(row['学号']): row['姓名'] for _, row in df.iterrows()}

            # 检查是否为文件夹项目
            is_folder_project = rename_format.get('is_folder', False)

            # 收集已交作业学生
            submitted_files = self._collect_submitted_files(
                homework_dir, all_students, id_to_name, is_folder_project, log_callback
            )

            # 处理未交作业名单
            self._process_missing_students(df, submitted_files, homework_dir, output_dir, log_callback)

            # 处理重复提交名单
            self._process_repeated_submissions(df, submitted_files, homework_dir, output_dir, log_callback)

            # 重命名文件
            rename_count = self.file_renamer.rename_files(
                df, homework_dir, rename_format, log_callback
            )
            self._log(f"成功重命名 {rename_count} 个学生的文件。", log_callback)

            self._log(f"\n{'-'*50}", log_callback)
            self._log(f"{project_name} 项目处理完成", log_callback)
            self._log(f"{'-'*50}\n", log_callback)

        except Exception as e:
            self._log(f"处理失败：{str(e)}", log_callback)
            raise

    def rename_files_only(self, roster_path: str, homework_dir: str, 
                         rename_format: dict, log_callback: Optional[Callable] = None) -> int:
        """
        仅重命名文件
        """
        try:
            df = self._read_roster(roster_path)
            count = self.file_renamer.rename_files(df, homework_dir, rename_format, log_callback)
            return count
        except Exception as e:
            self._log(f"重命名失败：{str(e)}", log_callback)
            raise

    def batch_check_submissions(self, roster_path: str, parent_dir: str,
                          rename_format: dict = None, 
                          selected_folders: list = None,
                          log_callback: Optional[Callable] = None) -> str:
        """
        批量检查多个子文件夹的提交情况并生成汇总报告
        :param roster_path: 花名册路径
        :param parent_dir: 母文件夹路径（包含多个实验子文件夹）
        :param rename_format: 重命名格式配置（可选，为None则不重命名）
        :param selected_folders: 指定要扫描的子文件夹列表（None则扫描全部）
        :param log_callback: 日志回调函数
        :return: 生成的Excel报告路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        import datetime
        
        self._log(f"📂 开始扫描母文件夹: {parent_dir}", log_callback)
        
        # 1. 读取花名册
        df_roster = self._read_roster(roster_path)
        all_students = set(df_roster['姓名'].tolist())
        id_to_name = {str(row['学号']): row['姓名'] for _, row in df_roster.iterrows()}
        
        # 2. 获取所有子文件夹（排除系统文件夹）
        all_subfolders = []
        for item in os.listdir(parent_dir):
            item_path = os.path.join(parent_dir, item)
            if os.path.isdir(item_path) and not item.startswith('.'):
                all_subfolders.append(item)
        
        # 2. 处理子文件夹顺序 - 直接使用 selected_folders 的顺序，不进行额外排序
        if selected_folders:
            # 验证指定的文件夹是否都存在
            valid_folders = []
            invalid_folders = []
            for folder in selected_folders:
                folder_path = os.path.join(parent_dir, folder)
                if os.path.exists(folder_path) and os.path.isdir(folder_path):
                    valid_folders.append(folder)
                else:
                    invalid_folders.append(folder)
            
            if invalid_folders:
                self._log(f"⚠️  警告: 以下文件夹不存在，已忽略: {', '.join(invalid_folders)}", log_callback)
            
            # 关键：直接使用 selected_folders 中的顺序，不再排序
            subfolders = [f for f in selected_folders if f in valid_folders]
            self._log(f"使用指定的 {len(subfolders)} 个子文件夹 (按保存的配置顺序)", log_callback)
            self._log(f"文件夹顺序: {', '.join(subfolders)}", log_callback)
        else:
            # 如果没有指定，按名称排序
            all_subfolders.sort()
            subfolders = all_subfolders
            self._log(f"使用全部 {len(subfolders)} 个子文件夹 (按名称排序)", log_callback)
        
        if not subfolders:
            raise Exception("母文件夹下没有找到有效的子文件夹（实验目录）")
        
        # 3. 为每个学生初始化一个字典，记录每个实验的状态
        student_status = {}
        for _, row in df_roster.iterrows():
            student_id = str(row['学号'])
            student_name = row['姓名']
            student_status[student_name] = {
                '学号': student_id,
                '姓名': student_name,
                **{folder: '未交' for folder in subfolders}  # 默认所有实验都未交
            }
        
        # 4. 遍历每个子文件夹，检查提交情况（可选重命名）
        for folder in subfolders:
            folder_path = os.path.join(parent_dir, folder)
            self._log(f"\n--- 检查子文件夹: {folder} ---", log_callback)
            
            # 收集此文件夹中已提交的学生
            submitted_files = self._collect_submitted_files(
                folder_path, all_students, id_to_name, False, None  # 最后一个参数设为None，不记录日志细节
            )
            
            # 更新状态
            for student_name in submitted_files.keys():
                if student_name in student_status:
                    student_status[student_name][folder] = '已交'
            
            self._log(f"  已交: {len(submitted_files)}人", log_callback)
            
            # 可选：执行重命名
            if rename_format:
                rename_count = self.file_renamer.rename_files(
                    df_roster, folder_path, rename_format, None
                )
                self._log(f"  重命名: {rename_count}个文件", log_callback)
        
        # 5. 构建汇总DataFrame
        # 列顺序：学号、姓名、实验1、实验2...
        columns = ['学号', '姓名'] + subfolders
        data = []
        
        for student_name, status_dict in student_status.items():
            row = [status_dict['学号'], status_dict['姓名']]
            row.extend([status_dict[folder] for folder in subfolders])
            data.append(row)
        
        df_summary = pd.DataFrame(data, columns=columns)
        
        # 6. 生成Excel报告（使用openpyxl以便设置单元格样式）
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        parent_folder_name = os.path.basename(parent_dir.rstrip(os.sep))
        output_filename = f"作业提交汇总_{parent_folder_name}_{timestamp}.xlsx"
        output_dir = os.path.join(parent_dir, "作业汇总报告")
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        output_path = os.path.join(output_dir, output_filename)
        
        # 使用openpyxl创建带格式的Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, index=False, sheet_name='提交汇总')
            
            # 获取workbook和worksheet对象以设置格式
            workbook = writer.book
            worksheet = writer.sheets['提交汇总']
            
            # 定义红色填充（用于“未交”单元格）
            red_fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')
            
            # 遍历所有单元格，为“未交”标记红色
            for row in worksheet.iter_rows(min_row=2, max_row=len(df_summary)+1, min_col=3, max_col=len(columns)):
                for cell in row:
                    if cell.value == '未交':
                        cell.fill = red_fill
            
            # 设置列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 冻结前两列（学号、姓名）
            worksheet.freeze_panes = 'C2'
        
        # 7. 统计信息
        total_students = len(df_summary)
        total_labs = len(subfolders)
        total_submissions = (df_summary.iloc[:, 2:] == '已交').sum().sum()  # 统计所有“已交”
        submission_rate = total_submissions / (total_students * total_labs) * 100 if total_students * total_labs > 0 else 0
        
        self._log(f"\n" + "="*60, log_callback)
        self._log(f"📊 汇总统计:", log_callback)
        self._log(f"  学生总数: {total_students}", log_callback)
        self._log(f"  实验总数: {total_labs}", log_callback)
        self._log(f"  总提交次数: {total_submissions}", log_callback)
        self._log(f"  总提交率: {submission_rate:.1f}%", log_callback)
        self._log(f"  报告位置: {output_path}", log_callback)
        self._log("="*60, log_callback)
        
        return output_path

    def _read_roster(self, roster_path: str) -> pd.DataFrame:
        """读取花名册文件"""
        return pd.read_excel(roster_path, dtype={'学号': str})

    def _collect_submitted_files(self, homework_dir: str, all_students: set, 
                               id_to_name: Dict[str, str], is_folder_project: bool,
                               log_callback: Optional[Callable]) -> Dict[str, List[str]]:
        """收集已提交作业的学生和文件"""
        submitted_files = {}

        if not os.path.exists(homework_dir):
            self._log(f"警告：作业文件夹不存在: {homework_dir}", log_callback)
            return submitted_files

        if is_folder_project:
            # 处理文件夹项目
            for item in os.listdir(homework_dir):
                item_path = os.path.join(homework_dir, item)
                if os.path.isdir(item_path):
                    self._match_student(item, item, all_students, id_to_name, submitted_files)
        else:
            # 处理文件项目
            for filename in os.listdir(homework_dir):
                filepath = os.path.join(homework_dir, filename)
                if filename.startswith('~$') or os.path.isdir(filepath):
                    continue
                self._match_student(filename, filename, all_students, id_to_name, submitted_files)

        return submitted_files

    def _match_student(self, search_text: str, file_item: str, all_students: set,
                      id_to_name: Dict[str, str], submitted_files: Dict[str, List[str]]):
        """匹配学生姓名或学号"""
        # 先尝试匹配姓名
        for name in all_students:
            if name in search_text:
                if name not in submitted_files:
                    submitted_files[name] = []
                submitted_files[name].append(file_item)
                return

        # 再尝试匹配学号
        for student_id, name in id_to_name.items():
            if student_id in search_text:
                if name not in submitted_files:
                    submitted_files[name] = []
                submitted_files[name].append(file_item)
                return

    def _process_missing_students(self, df: pd.DataFrame, submitted_files: Dict[str, List[str]],
                                homework_dir: str, output_dir: str, log_callback: Optional[Callable]):
        """处理未交作业学生"""
        submitted_students = set(submitted_files.keys())
        all_students = set(df['姓名'].tolist())
        missing_students = all_students - submitted_students

        if missing_students:
            missing_df = df[df['姓名'].isin(missing_students)].copy()
            missing_df['学号'] = missing_df['学号'].astype(str)

            folder_name = os.path.basename(homework_dir.rstrip(os.sep))
            output_path = os.path.join(output_dir, f"未交作业名单_{folder_name}.xlsx")
            missing_df.to_excel(output_path, index=False)

            self._log(f"生成未交报告：{output_path}", log_callback)
            self._log(f"未交人数：{len(missing_students)}，名单：{', '.join(missing_students)}", log_callback)
        else:
            self._log("所有学生均已提交作业！", log_callback)

    def _process_repeated_submissions(self, df: pd.DataFrame, submitted_files: Dict[str, List[str]],
                                    homework_dir: str, output_dir: str, log_callback: Optional[Callable]):
        """处理重复提交"""
        repeated_records = []
        for name, files in submitted_files.items():
            if len(files) > 1:
                student_info = df[df['姓名'] == name].iloc[0]
                marked_files = [f"*{f}" for f in files]
                repeated_records.append({
                    "学号": student_info['学号'],
                    "姓名": name,
                    "提交文件": ", ".join(marked_files),
                    "提交次数": len(files)
                })

        if repeated_records:
            repeat_df = pd.DataFrame(repeated_records)
            folder_name = os.path.basename(homework_dir.rstrip(os.sep))
            repeat_path = os.path.join(output_dir, f"重复提交名单_{folder_name}.xlsx")
            repeat_df.to_excel(repeat_path, index=False)

            self._log(f"生成重复提交报告：{repeat_path}", log_callback)
            self._log(f"重复提交人数：{len(repeated_records)}，名单：{', '.join([r['姓名'] for r in repeated_records])}", log_callback)
        else:
            self._log("没有重复提交的学生。", log_callback)

    def _log(self, message: str, log_callback: Optional[Callable]):
        """记录日志"""
        if log_callback:
            log_callback(message)
        else:
            print(message)